from __future__ import annotations

import csv
import ipaddress
import os
import re
from dataclasses import dataclass
from pathlib import Path
from tkinter import BOTH, END, LEFT, RIGHT, X, StringVar, Tk, filedialog, messagebox
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
from urllib.parse import urlparse

from openpyxl import load_workbook
from app_icon import apply_app_icon


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENSCAN_DIR = PROJECT_ROOT / "tools" / "enscan"
DOMAINTOIP_DIR = PROJECT_ROOT / "tools" / "domaintoIP"
DEFAULT_COMPANY_FILE = ENSCAN_DIR / "gongsi.txt"
XLSX_PATTERNS = (
    "gongsi.txt批量查询任务结果-*.xlsx",
    "outs/gongsi.txt批量查询任务结果-*.xlsx",
)

IPV4_RE = re.compile(
    r"(?<!\d)(?:25[0-5]|2[0-4]\d|1\d\d|[1-9]?\d)"
    r"(?:\.(?:25[0-5]|2[0-4]\d|1\d\d|[1-9]?\d)){3}(?!\d)"
)
URL_RE = re.compile(r"\b[a-zA-Z][a-zA-Z0-9+.-]*://[^\s]+")
DOMAIN_RE = re.compile(
    r"(?<![a-zA-Z0-9_-])(?:\*\.)?"
    r"(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)+"
    r"[a-zA-Z]{2,63}(?![a-zA-Z0-9_-])"
)
READ_ENCODINGS = ("utf-8", "utf-8-sig", "gb18030", "gbk", "latin-1")
INVISIBLE_CHARS = "\ufeff\u200b\u200c\u200d\u2060"

MULTI_LEVEL_SUFFIXES = {
    "ac.cn", "com.cn", "edu.cn", "gov.cn", "mil.cn", "net.cn", "org.cn",
    "ah.cn", "bj.cn", "cq.cn", "fj.cn", "gd.cn", "gs.cn", "gx.cn", "gz.cn",
    "ha.cn", "hb.cn", "he.cn", "hi.cn", "hk.cn", "hl.cn", "hn.cn", "jl.cn",
    "js.cn", "jx.cn", "ln.cn", "mo.cn", "nm.cn", "nx.cn", "qh.cn", "sc.cn",
    "sd.cn", "sh.cn", "sn.cn", "sx.cn", "tj.cn", "tw.cn", "xj.cn", "xz.cn",
    "yn.cn", "zj.cn", "co.uk", "ac.uk", "gov.uk", "ltd.uk", "me.uk", "net.uk",
    "org.uk", "plc.uk", "sch.uk", "com.au", "edu.au", "gov.au", "id.au", "net.au",
    "org.au", "asn.au", "com.hk", "edu.hk", "gov.hk", "idv.hk", "net.hk", "org.hk",
    "com.tw", "edu.tw", "gov.tw", "idv.tw", "net.tw", "org.tw", "co.jp", "ac.jp",
    "ad.jp", "ed.jp", "go.jp", "gr.jp", "lg.jp", "ne.jp", "or.jp", "com.sg",
    "edu.sg", "gov.sg", "net.sg", "org.sg", "per.sg", "com.my", "edu.my", "gov.my",
    "net.my", "org.my", "co.kr", "ne.kr", "or.kr", "re.kr", "pe.kr", "go.kr",
    "mil.kr", "ac.kr", "co.nz", "ac.nz", "geek.nz", "gen.nz", "govt.nz", "health.nz",
    "iwi.nz", "kiwi.nz", "maori.nz", "mil.nz", "net.nz", "org.nz", "school.nz",
}
COMMON_COUNTRY_SECOND_LEVELS = {
    "ac", "co", "com", "edu", "gov", "mil", "net", "nom", "org", "sch",
}

MODE_CONFIGS = {
    "keep": {
        "title": "保留模式",
        "desc": "提取命中单位对应的 URL，输出 company_url_ok.txt。",
        "primary_file": "company_url_ok.txt",
        "primary_label": "保留URL",
        "primary_attr": "kept_url_lines",
    },
    "reverse": {
        "title": "反向模式",
        "desc": "去除命中单位对应的 URL，输出 company_url_ok.txt。",
        "primary_file": "company_url_ok.txt",
        "primary_label": "反向保留URL",
        "primary_attr": "removed_url_lines",
    },
}


@dataclass(frozen=True)
class InputTarget:
    mode: str
    value: str


@dataclass
class XlsxMatchRecord:
    sheet_name: str
    row_index: int
    matched_company: str
    source_value: str
    extracted_targets: list[str]
    row_values: list[str]


@dataclass
class BatchFilterContext:
    company_names: list[str]
    extracted_targets: list[InputTarget]
    xlsx_match_records: list[XlsxMatchRecord]
    matched_result_lines: list[str]
    matched_root_domains: set[str]
    matched_domains: set[str]
    matched_ips: set[str]
    domain_related_ips: set[str]
    kept_url_lines: list[str]
    removed_url_lines: list[str]


def read_text_auto(path: Path) -> str:
    data = path.read_bytes()
    for encoding in READ_ENCODINGS:
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def write_lines(path: Path, lines: list[str]) -> None:
    text = "\n".join(lines)
    if lines:
        text += "\n"
    path.write_text(text, encoding="utf-8")


def write_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.writer(fp)
        writer.writerow(header)
        writer.writerows(rows)


def strip_invisible_edges(value) -> str:
    if value is None:
        return ""
    return str(value).translate({ord(ch): None for ch in INVISIBLE_CHARS}).strip()


def normalize_for_match(text: str) -> str:
    text = strip_invisible_edges(text)
    if not text:
        return ""
    return re.sub(r"\s+", "", text).lower()


def is_ipv4(value: str) -> bool:
    try:
        ipaddress.IPv4Address(value)
        return True
    except ValueError:
        return False


def normalize_host_or_ip(value: str) -> str:
    text = value.strip().strip('"\'').strip()
    if not text:
        return ""

    if "://" in text:
        parsed = urlparse(text)
        if parsed.hostname:
            return parsed.hostname.lower().rstrip(".")

    if text.startswith("//"):
        parsed = urlparse(f"http:{text}")
        if parsed.hostname:
            return parsed.hostname.lower().rstrip(".")

    text = text.split("/", 1)[0]
    text = text.rsplit("@", 1)[-1]
    text = text.strip("[](){}<>.,;\"' ").rstrip(".").lower()
    if not text:
        return ""

    if text.count(":") == 1:
        host, port = text.rsplit(":", 1)
        if port.isdigit():
            text = host

    return text


def get_registrable_domain(host: str) -> str:
    host = normalize_host_or_ip(host)
    if not host or is_ipv4(host):
        return host

    labels = [part for part in host.split(".") if part]
    if len(labels) <= 2:
        return ".".join(labels)

    last_two = ".".join(labels[-2:])
    if last_two in MULTI_LEVEL_SUFFIXES and len(labels) >= 3:
        return ".".join(labels[-3:])

    if len(labels[-1]) == 2 and labels[-2] in COMMON_COUNTRY_SECOND_LEVELS:
        return ".".join(labels[-3:])

    return ".".join(labels[-2:])


def normalize_root_domain_or_ip(value: str) -> InputTarget:
    normalized = normalize_host_or_ip(value)
    if not normalized:
        raise ValueError("输入为空，请删掉空行后重试。")

    if is_ipv4(normalized):
        return InputTarget(mode="ip", value=normalized)

    if "." not in normalized:
        raise ValueError(f"不是有效根域名或 IP: {value}")

    return InputTarget(mode="domain", value=get_registrable_domain(normalized))


def dedupe_input_targets(values: list[str]) -> list[InputTarget]:
    results: list[InputTarget] = []
    seen: set[tuple[str, str]] = set()

    for value in values:
        item = normalize_root_domain_or_ip(value)
        key = (item.mode, item.value)
        if key not in seen:
            seen.add(key)
            results.append(item)

    return results


def extract_targets_from_text(text: str) -> list[str]:
    text = strip_invisible_edges(text)
    if not text:
        return []

    seen: set[str] = set()
    results: list[str] = []

    def add(value: str) -> None:
        normalized = normalize_host_or_ip(value)
        if not normalized:
            return
        if is_ipv4(normalized):
            target = normalized
        else:
            if "." not in normalized:
                return
            target = get_registrable_domain(normalized)
        if target not in seen:
            seen.add(target)
            results.append(target)

    add(text)
    for match in URL_RE.findall(text):
        add(match)
    for match in IPV4_RE.findall(text):
        add(match)
    for match in DOMAIN_RE.findall(text):
        add(match.lstrip("*."))

    return results


def parse_company_names(text: str) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()

    for line in text.replace(",", "\n").splitlines():
        raw = strip_invisible_edges(line)
        if not raw:
            continue
        normalized = normalize_for_match(raw)
        if normalized and normalized not in seen:
            seen.add(normalized)
            names.append(raw)

    if not names:
        raise ValueError("请输入单位全称或备案单位名称，一行一个。")

    return names


def find_latest_xlsx(base_dir: Path) -> Path | None:
    candidates = []
    for pattern in XLSX_PATTERNS:
        candidates.extend(base_dir.glob(pattern))
    files = [item for item in candidates if item.is_file()]
    if not files:
        return None
    return max(files, key=lambda item: item.stat().st_mtime)


def choose_source_value(row_values: list[str]) -> str:
    for index in (2, 1, 0):
        if index < len(row_values):
            value = row_values[index]
            if extract_targets_from_text(value):
                return value
    return " ".join(row_values)


def parse_xlsx_targets(company_text: str, xlsx_path: Path) -> tuple[list[str], list[XlsxMatchRecord]]:
    company_names = parse_company_names(company_text)
    company_map = {normalize_for_match(name): name for name in company_names}
    matched_targets: list[str] = []
    match_records: list[XlsxMatchRecord] = []
    seen_targets: set[str] = set()

    workbook = load_workbook(xlsx_path, read_only=False, data_only=True, keep_links=False)
    try:
        preferred_sheets = sorted(
            workbook.worksheets,
            key=lambda sheet: (0 if "ICP备案" in sheet.title else 1, sheet.title),
        )

        for sheet in preferred_sheets:
            sheet_targets_before = len(matched_targets)

            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index == 1:
                    continue

                row_values = [strip_invisible_edges(value) for value in row if strip_invisible_edges(value)]
                if not row_values:
                    continue

                row_text = normalize_for_match(" ".join(row_values))
                if not row_text:
                    continue

                matched_company = ""
                for normalized, original in company_map.items():
                    if normalized in row_text:
                        matched_company = original
                        break

                if not matched_company:
                    continue

                source_value = choose_source_value(row_values)
                extracted_targets = extract_targets_from_text(source_value)
                if not extracted_targets:
                    continue

                match_records.append(
                    XlsxMatchRecord(
                        sheet_name=sheet.title,
                        row_index=row_index,
                        matched_company=matched_company,
                        source_value=source_value,
                        extracted_targets=extracted_targets,
                        row_values=row_values,
                    )
                )

                for item in extracted_targets:
                    if item not in seen_targets:
                        seen_targets.add(item)
                        matched_targets.append(item)

            if sheet.title.find("ICP备案") >= 0 and len(matched_targets) > sheet_targets_before:
                break

            if len(matched_targets) > sheet_targets_before:
                break
    finally:
        workbook.close()

    if not matched_targets:
        raise ValueError("未从 Enscan XLSX 中按单位名称提取到有效根域名或 IP。")

    return matched_targets, match_records


def extract_ipv4s(text: str) -> set[str]:
    return set(IPV4_RE.findall(text))


def extract_domains(text: str) -> set[str]:
    domains: set[str] = set()

    for url in URL_RE.findall(text):
        host = normalize_host_or_ip(url)
        if host and not is_ipv4(host):
            domains.add(host)

    for match in DOMAIN_RE.findall(text):
        host = normalize_host_or_ip(match.lstrip("*."))
        if host and not is_ipv4(host):
            domains.add(host)

    return domains


def should_keep_url_line(
    line: str,
    domain_targets: set[str],
    ip_targets: set[str],
    domain_related_ips: set[str],
) -> bool:
    host = normalize_host_or_ip(line)
    if not host:
        return False

    if is_ipv4(host):
        return host in ip_targets or host in domain_related_ips

    return get_registrable_domain(host) in domain_targets


def analyze_files(company_text: str, xlsx_path: Path, result_path: Path, url_path: Path) -> BatchFilterContext:
    company_names = parse_company_names(company_text)
    xlsx_targets, match_records = parse_xlsx_targets(company_text, xlsx_path)
    targets = dedupe_input_targets(xlsx_targets)

    result_lines = [line.strip() for line in read_text_auto(result_path).splitlines() if line.strip()]
    url_lines = [line.strip() for line in read_text_auto(url_path).splitlines() if line.strip()]

    matched_result_lines: list[str] = []
    matched_root_domains: set[str] = set()
    matched_domains: set[str] = set()
    matched_ips: set[str] = set()
    domain_related_ips: set[str] = set()

    domain_targets = {item.value for item in targets if item.mode == "domain"}
    ip_targets = {item.value for item in targets if item.mode == "ip"}

    for line in result_lines:
        ips = extract_ipv4s(line)
        domains = extract_domains(line)
        roots = {get_registrable_domain(domain) for domain in domains if domain}

        matched_by_ip = bool(ip_targets.intersection(ips))
        matched_by_domain = bool(domain_targets.intersection(roots))
        if not (matched_by_ip or matched_by_domain):
            continue

        matched_result_lines.append(line)
        matched_ips.update(ips)
        matched_domains.update(domains)
        matched_root_domains.update(roots)
        if matched_by_domain:
            domain_related_ips.update(ips)

    kept_url_lines: list[str] = []
    removed_url_lines: list[str] = []

    for line in url_lines:
        if should_keep_url_line(line, domain_targets, ip_targets, domain_related_ips):
            kept_url_lines.append(line)
        else:
            removed_url_lines.append(line)

    return BatchFilterContext(
        company_names=company_names,
        extracted_targets=targets,
        xlsx_match_records=match_records,
        matched_result_lines=matched_result_lines,
        matched_root_domains=matched_root_domains,
        matched_domains=matched_domains,
        matched_ips=matched_ips,
        domain_related_ips=domain_related_ips,
        kept_url_lines=kept_url_lines,
        removed_url_lines=removed_url_lines,
    )


def save_outputs(context: BatchFilterContext, output_dir: Path, mode_key: str) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    config = MODE_CONFIGS[mode_key]

    primary_path = output_dir / config["primary_file"]
    primary_lines = list(getattr(context, config["primary_attr"]))
    write_lines(primary_path, primary_lines)
    return {"primary": primary_path}


def latest_file(base: Path, pattern: str) -> Path | None:
    files = [item for item in base.glob(pattern) if item.is_file()]
    if not files:
        return None
    return max(files, key=lambda item: item.stat().st_mtime)


def latest_file_by_patterns(base: Path, patterns: list[str]) -> Path | None:
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(item for item in base.glob(pattern) if item.is_file())
    if not candidates:
        return None
    return max(candidates, key=lambda item: item.stat().st_mtime)


class CompanyFullnameUrlFilterGui:
    def __init__(self, root: Tk) -> None:
        self.root = root
        apply_app_icon(self.root, app_id="ziweiqi.desktop.company-filter")
        self.root.title("By：ziweiqi")
        self.root.geometry("1180x800")
        self.root.minsize(1080, 720)

        self.workdir = PROJECT_ROOT
        self.last_output_path: Path | None = None

        self.xlsx_file_var = StringVar(value=str(self._default_xlsx_file()))
        self.result_file_var = StringVar(value=str(self._default_result_file()))
        self.url_file_var = StringVar(value=str(self._default_url_file()))
        self.output_dir_var = StringVar(value=str(PROJECT_ROOT))
        self.mode_var = StringVar(value="keep")
        self.summary_var = StringVar(value="等待开始。")
        self._apply_env_defaults()

        self._build_ui()
        self._center_window()

    def _apply_env_defaults(self) -> None:
        xlsx_file = os.environ.get("ZIWEIQI_XLSX_FILE", "").strip()
        result_file = os.environ.get("ZIWEIQI_RESULT_FILE", "").strip()
        url_file = os.environ.get("ZIWEIQI_URL_FILE", "").strip()
        output_dir = os.environ.get("ZIWEIQI_OUTPUT_DIR", "").strip()
        if xlsx_file:
            self.xlsx_file_var.set(xlsx_file)
        if result_file:
            self.result_file_var.set(result_file)
        if url_file:
            self.url_file_var.set(url_file)
        if output_dir:
            self.output_dir_var.set(output_dir)

    def _default_xlsx_file(self) -> Path:
        latest = find_latest_xlsx(ENSCAN_DIR)
        return latest if latest else ENSCAN_DIR / "outs"

    def _default_result_file(self) -> Path:
        candidates = [
            DOMAINTOIP_DIR / "result.txt",
            PROJECT_ROOT / "result.txt",
            PROJECT_ROOT / "results.txt",
        ]
        for path in candidates:
            if path.exists():
                return path
        return DOMAINTOIP_DIR / "result.txt"

    def _default_url_file(self) -> Path:
        latest_asset_txt = latest_file_by_patterns(
            PROJECT_ROOT / "results",
            ["*资产已去重_*.txt", "*_results2.txt"],
        )
        if latest_asset_txt:
            return latest_asset_txt
        candidates = [
            PROJECT_ROOT / "url.txt",
            PROJECT_ROOT / "results" / "url.txt",
        ]
        for path in candidates:
            if path.exists():
                return path
        return PROJECT_ROOT / "url.txt"

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill=BOTH, expand=True)

        title = ttk.Label(main, text="按单位全称提取及剔除工具", font=("Microsoft YaHei UI", 14, "bold"))
        title.pack(anchor="w")

        desc = ttk.Label(
            main,
            text=(
                "输入单位全称或备案单位名，一行一个。程序会先从 Enscan XLSX 中提取对应根域名/IP，"
                "再到 domaintoIP 下的 result.txt 中补齐命中的子域名和 IP，最后对 results 目录下的 资产已去重_*.txt 执行提取或去除。"
            ),
            wraplength=1140,
            justify="left",
        )
        desc.pack(anchor="w", pady=(6, 12))

        target_frame = ttk.LabelFrame(main, text="批量输入目标单位全称", padding=10)
        target_frame.pack(fill=BOTH, pady=(0, 8))

        tip = ttk.Label(
            target_frame,
            text="支持单位全称或备案单位名称，一行一个。匹配时会自动忽略空格和大小写差异。",
            wraplength=1120,
            justify="left",
        )
        tip.pack(anchor="w")

        target_btns = ttk.Frame(target_frame)
        target_btns.pack(fill=X, pady=(8, 6))
        ttk.Button(target_btns, text="导入目标文件", command=self._import_targets).pack(side=LEFT)
        ttk.Button(target_btns, text="清空输入", command=self._clear_targets).pack(side=LEFT, padx=8)
        ttk.Button(target_btns, text="使用示例", command=self._show_usage_example).pack(side=LEFT)

        self.target_text = ScrolledText(target_frame, height=10)
        self.target_text.pack(fill=BOTH, expand=True)

        self._build_path_selector(main, "Enscan XLSX 文件", self.xlsx_file_var, self._choose_xlsx_file)
        self._build_path_selector(main, "domaintoIP 下的result.txt 子域名及对应ip文件", self.result_file_var, self._choose_result_file)
        self._build_path_selector(main, "ziweiqi\\results 目录下 资产已去重_*.txt", self.url_file_var, self._choose_url_file)
        self._build_path_selector(main, "输出目录", self.output_dir_var, self._choose_output_dir, choose_dir=True)

        mode_frame = ttk.LabelFrame(main, text="模式", padding=8)
        mode_frame.pack(fill=X, pady=(0, 8))
        ttk.Radiobutton(mode_frame, text=MODE_CONFIGS["keep"]["title"], value="keep", variable=self.mode_var).pack(side=LEFT)
        ttk.Radiobutton(mode_frame, text=MODE_CONFIGS["reverse"]["title"], value="reverse", variable=self.mode_var).pack(side=LEFT, padx=12)
        ttk.Label(mode_frame, textvariable=self.summary_var).pack(side=RIGHT)

        action_frame = ttk.Frame(main)
        action_frame.pack(fill=X, pady=(0, 10))
        ttk.Button(action_frame, text="开始分析并导出", command=self.on_analyze).pack(side=LEFT)
        ttk.Button(action_frame, text="打开输出目录", command=self._open_output_dir).pack(side=LEFT, padx=8)
        ttk.Button(action_frame, text="清空日志", command=self.clear_log).pack(side=LEFT, padx=8)
        ttk.Label(action_frame, textvariable=self.summary_var).pack(side=RIGHT)

        log_frame = ttk.LabelFrame(main, text="运行日志 / 结果预览", padding=8)
        log_frame.pack(fill=BOTH, expand=True)
        self.log_text = ScrolledText(log_frame, height=28)
        self.log_text.pack(fill=BOTH, expand=True)
        self.log_text.configure(state="disabled")

    def _center_window(self) -> None:
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = max(0, (screen_width - width) // 2)
        y = max(20, (screen_height - height) // 6)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def _build_path_selector(
        self,
        parent: ttk.Frame,
        title: str,
        variable: StringVar,
        command,
        choose_dir: bool = False,
    ) -> None:
        frame = ttk.LabelFrame(parent, text=title, padding=8)
        frame.pack(fill=X, pady=(0, 8))
        ttk.Entry(frame, textvariable=variable).pack(side=LEFT, fill=X, expand=True, padx=(0, 8))
        ttk.Button(frame, text="选择目录" if choose_dir else "选择文件", command=command).pack(side=LEFT)

    def log(self, message: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert(END, message + "\n")
        self.log_text.see(END)
        self.log_text.configure(state="disabled")

    def clear_log(self) -> None:
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", END)
        self.log_text.configure(state="disabled")
        self.summary_var.set("日志已清空。")

    def _load_default_company_content(self) -> None:
        if not DEFAULT_COMPANY_FILE.exists():
            return
        content = read_text_auto(DEFAULT_COMPANY_FILE)
        self.target_text.delete("1.0", END)
        self.target_text.insert("1.0", content)
        line_count = len([line for line in content.splitlines() if line.strip()])
        self.summary_var.set(f"已加载默认单位名单 {line_count} 条。")

    def _import_targets(self) -> None:
        path = filedialog.askopenfilename(
            title="选择目标列表文件",
            initialdir=self._safe_initial_dir(str(DEFAULT_COMPANY_FILE)),
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")],
        )
        if not path:
            return

        content = read_text_auto(Path(path))
        self.target_text.delete("1.0", END)
        self.target_text.insert("1.0", content)
        line_count = len([line for line in content.splitlines() if line.strip()])
        self.summary_var.set(f"已导入 {line_count} 条目标单位。")

    def _clear_targets(self) -> None:
        self.target_text.delete("1.0", END)
        self.summary_var.set("输入已清空。")

    def _show_usage_example(self) -> None:
        example_text = (
            "湖南大学\n"
            "国家林业和草原局中南调查规划院\n"
            "湖南建设投资集团有限责任公司\n"
        )
        self.target_text.delete("1.0", END)
        self.target_text.insert("1.0", example_text)
        self.summary_var.set("已填充使用示例。")

    def _choose_xlsx_file(self) -> None:
        path = filedialog.askopenfilename(
            title="选择 Enscan XLSX",
            initialdir=self._safe_initial_dir(self.xlsx_file_var.get()),
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
        )
        if path:
            self.xlsx_file_var.set(path)

    def _choose_result_file(self) -> None:
        path = filedialog.askopenfilename(
            title="选择 domaintoIP 下的 results.txt",
            initialdir=self._safe_initial_dir(self.result_file_var.get()),
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")],
        )
        if path:
            self.result_file_var.set(path)

    def _choose_url_file(self) -> None:
        path = filedialog.askopenfilename(
            title="选择 URL 结果文件",
            initialdir=self._safe_initial_dir(self.url_file_var.get()),
            filetypes=[("Text Files", "*.txt"), ("CSV Files", "*.csv"), ("All Files", "*.*")],
        )
        if path:
            self.url_file_var.set(path)

    def _choose_output_dir(self) -> None:
        path = filedialog.askdirectory(
            title="选择输出目录",
            initialdir=self._safe_initial_dir(self.output_dir_var.get()),
        )
        if path:
            self.output_dir_var.set(path)

    def _open_output_dir(self) -> None:
        output_dir = Path(self.output_dir_var.get()).expanduser()
        if not output_dir.exists():
            messagebox.showerror("打开失败", f"输出目录不存在: {output_dir}")
            return
        try:
            os.startfile(output_dir)  # type: ignore[attr-defined]
        except OSError as exc:
            messagebox.showerror("打开失败", str(exc))

    def _safe_initial_dir(self, raw_path: str) -> str:
        path = Path(raw_path).expanduser()
        if path.exists():
            return str(path if path.is_dir() else path.parent)
        return str(self.workdir)

    def _collect_target_text(self) -> str:
        return self.target_text.get("1.0", END).strip()

    def on_analyze(self) -> None:
        try:
            xlsx_path, result_path, url_path, output_dir, company_text = self._collect_inputs()
            context = analyze_files(company_text, xlsx_path, result_path, url_path)
            mode_key = self._current_mode()
            output_paths = save_outputs(context, output_dir, mode_key)
        except Exception as exc:
            messagebox.showerror("分析失败", str(exc))
            self.summary_var.set("分析失败。")
            return

        self.last_output_path = output_paths["primary"]
        config = MODE_CONFIGS[self._current_mode()]
        self.log("")
        self.log("=" * 88)
        self.log(f"当前模式: {config['title']}")
        self.log(f"输入单位数量: {len(context.company_names)}")
        self.log(f"XLSX 提取目标数量: {len(context.extracted_targets)}")
        self.log(f"XLSX 命中行数: {len(context.xlsx_match_records)}")
        self.log(f"domaintoIP/results.txt 命中行数: {len(context.matched_result_lines)}")
        self.log(f"命中根域名数量: {len(context.matched_root_domains)}")
        self.log(f"命中子域名数量: {len(context.matched_domains)}")
        self.log(f"命中 IP 数量: {len(context.matched_ips)}")
        self.log(f"{config['primary_label']} 行数: {len(getattr(context, config['primary_attr']))}")
        self.log(f"提取靶标文件: {output_paths['targets']}")
        self.log(f"result 命中文件: {output_paths['results']}")
        self.log(f"XLSX 命中明细: {output_paths['rows']}")
        self.log(f"最终输出: {output_paths['primary']}")

        self.summary_var.set(f"完成: {config['primary_file']} = {len(getattr(context, config['primary_attr']))}")
        messagebox.showinfo("完成", f"结果已生成:\n{output_paths['primary']}")

    def _collect_inputs(self) -> tuple[Path, Path, Path, Path, str]:
        xlsx_path = Path(self.xlsx_file_var.get()).expanduser()
        result_path = Path(self.result_file_var.get()).expanduser()
        url_path = Path(self.url_file_var.get()).expanduser()
        output_dir = Path(self.output_dir_var.get()).expanduser()
        company_text = self._collect_target_text()

        if not company_text:
            raise ValueError("请输入单位全称或备案单位名称。")
        if not xlsx_path.is_file():
            raise FileNotFoundError(f"Enscan XLSX 不存在: {xlsx_path}")
        if not result_path.is_file():
            raise FileNotFoundError(f"domaintoIP 下的 results.txt 不存在: {result_path}")
        if not url_path.is_file():
            raise FileNotFoundError(f"URL 结果文件不存在: {url_path}")
        if output_dir.exists() and not output_dir.is_dir():
            raise NotADirectoryError(f"输出目录不是文件夹: {output_dir}")

        return xlsx_path, result_path, url_path, output_dir, company_text

    def _current_mode(self) -> str:
        mode = self.mode_var.get().strip()
        return mode if mode in MODE_CONFIGS else "keep"


def main() -> None:
    root = Tk()
    CompanyFullnameUrlFilterGui(root)
    root.mainloop()


if __name__ == "__main__":
    main()
