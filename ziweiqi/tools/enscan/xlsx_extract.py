from __future__ import annotations

import argparse
import re
from pathlib import Path
from urllib.parse import urlsplit

import tkinter as tk
from tkinter import filedialog

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
ENSCAN_DIR = Path(__file__).resolve().parent
DEFAULT_COMPANY_FILE = ENSCAN_DIR / "gongsi.txt"
DEFAULT_OUTPUT_FILE = PROJECT_ROOT / "scanIPAndDomain.txt"
XLSX_PATTERNS = (
    "gongsi.txt批量查询任务结果-*.xlsx",
    "outs/gongsi.txt批量查询任务结果-*.xlsx",
)

INVISIBLE_CHARS = "\ufeff\u200b\u200c\u200d\u2060"
DOMAIN_RE = re.compile(
    r"(?<![@\w-])(?:[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?\.)+"
    r"[A-Za-z]{2,63}(?![\w-])"
)
ASCII_DOMAIN_LABEL_RE = re.compile(r"[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?")
IPV4_RE = re.compile(r"(?<![\d.])(?:\d{1,3}\.){3}\d{1,3}(?![\d.])")
TARGET_HEADER_NAMES = {
    "域名",
    "根域名",
    "主域名",
    "domain",
    "domains",
    "rootdomain",
    "rootdomains",
}
FALLBACK_HEADER_NAMES = {
    "网址",
    "网站",
    "url",
    "urls",
    "链接",
    "link",
    "links",
}


def strip_invisible_edges(value) -> str:
    if value is None:
        return ""
    return str(value).translate({ord(ch): None for ch in INVISIBLE_CHARS}).strip()


def normalize_for_match(text: str) -> str:
    text = strip_invisible_edges(text)
    if not text:
        return ""
    return re.sub(r"\s+", "", text).lower()


def is_valid_ipv4(value: str) -> bool:
    parts = value.split(".")
    if len(parts) != 4:
        return False
    try:
        return all(0 <= int(part) <= 255 for part in parts)
    except ValueError:
        return False


def target_sort_key(value: str) -> tuple:
    if is_valid_ipv4(value):
        return (0, tuple(int(part) for part in value.split(".")))
    return (1, value.lower())


def sort_targets(targets: list[str]) -> list[str]:
    return sorted(targets, key=target_sort_key)


def is_valid_domain(value: str) -> bool:
    if not value or "." not in value or len(value) > 253:
        return False
    try:
        value.encode("ascii")
    except UnicodeEncodeError:
        return False

    labels = value.rstrip(".").split(".")
    if len(labels) < 2 or any(not label for label in labels):
        return False

    ascii_labels = []
    for label in labels:
        ascii_label = label.lower()
        if len(ascii_label) > 63 or not ASCII_DOMAIN_LABEL_RE.fullmatch(ascii_label):
            return False
        ascii_labels.append(ascii_label)

    tld = ascii_labels[-1]
    if not (re.fullmatch(r"[a-z]{2,63}", tld) or tld.startswith("xn--")):
        return False

    return True


def normalize_target(value: str) -> str | None:
    candidate = strip_invisible_edges(value).strip("\"'[](){}<>;,，。；、")
    if not candidate:
        return None
    if "://" in candidate:
        try:
            candidate = urlsplit(candidate).hostname or ""
        except Exception:
            return None
    candidate = candidate.strip().lower().rstrip(".")
    if candidate.startswith("*."):
        candidate = candidate[2:]
    if candidate.startswith("www."):
        candidate = candidate[4:]
    if not candidate or "@" in candidate:
        return None
    if is_valid_ipv4(candidate):
        return candidate
    if is_valid_domain(candidate):
        return candidate
    return None


def extract_targets_from_text(text: str) -> list[str]:
    text = strip_invisible_edges(text)
    if not text:
        return []
    seen = set()
    results = []

    def add(value: str):
        normalized = normalize_target(value)
        if normalized and normalized not in seen:
            seen.add(normalized)
            results.append(normalized)

    add(text)
    for match in IPV4_RE.findall(text):
        add(match)
    for match in DOMAIN_RE.findall(text):
        add(match)
    return results


def load_company_names(company_file: Path) -> list[str]:
    if not company_file.exists():
        return []
    names = []
    for line in company_file.read_text(encoding="utf-8", errors="replace").splitlines():
        item = normalize_for_match(line)
        if item:
            names.append(item)
    return sorted(set(names), key=len, reverse=True)


def pick_sheets(workbook):
    sheets = workbook.worksheets
    if not sheets:
        raise RuntimeError("XLSX 没有任何 sheet")
    return sheets


def get_header_row(sheet) -> list[str]:
    try:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return []
    return [normalize_for_match(value) for value in first_row]


def find_header_index(headers: list[str], candidates: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in candidates:
            return index
    return None


def pick_target_columns(workbook) -> list[tuple[object, int, str]]:
    primary = []
    fallback = []

    for sheet in pick_sheets(workbook):
        headers = get_header_row(sheet)
        domain_index = find_header_index(headers, TARGET_HEADER_NAMES)
        if domain_index is not None:
            primary.append((sheet, domain_index, headers[domain_index] or "域名"))
            continue

        fallback_index = find_header_index(headers, FALLBACK_HEADER_NAMES)
        if fallback_index is not None:
            fallback.append((sheet, fallback_index, headers[fallback_index] or "网址/链接"))

    if primary:
        return primary
    if fallback:
        return fallback

    sheets = pick_sheets(workbook)
    return [(sheet, 2, "第三列") for sheet in sheets]


def find_latest_xlsx(base_dir: Path) -> Path | None:
    candidates = []
    for pattern in XLSX_PATTERNS:
        candidates.extend(base_dir.glob(pattern))
    files = [item for item in candidates if item.is_file()]
    if not files:
        return None
    return max(files, key=lambda item: item.stat().st_mtime)


def extract_targets_from_xlsx_file(xlsx_path: Path, company_file: Path, debug: bool = False) -> list[str]:
    company_names = load_company_names(company_file)
    workbook = load_workbook(xlsx_path, read_only=False, data_only=True, keep_links=False)
    results = []
    try:
        if debug:
            print(f"XLSX: {xlsx_path}")
            print(f"公司名单: {company_file} ({len(company_names)} 条)")
            print(f"sheet: {workbook.sheetnames}")

        target_columns = pick_target_columns(workbook)
        if debug:
            print("提取列: " + ", ".join(
                f"{sheet.title}.{column_name}(第{column_index + 1}列)"
                for sheet, column_index, column_name in target_columns
            ))

        for sheet, target_column_index, column_name in target_columns:
            if debug:
                print(f"\n扫描 sheet: {sheet.title}，目标列: {column_name}(第{target_column_index + 1}列)")

            sheet_hits = 0
            scanned_rows = 0
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index == 1:
                    continue
                scanned_rows += 1

                row_parts = []
                for value in row:
                    if value is None:
                        continue
                    part = strip_invisible_edges(value)
                    if part:
                        row_parts.append(part)
                if not row_parts:
                    continue

                row_text = normalize_for_match(" ".join(row_parts))
                if not row_text:
                    continue

                matched_company = None
                for company_name in company_names:
                    if company_name in row_text:
                        matched_company = company_name
                        break

                if company_names and not matched_company:
                    continue

                target_value = row[target_column_index] if len(row) > target_column_index else None
                if target_value is None:
                    continue

                candidates = extract_targets_from_text(str(target_value))
                if not candidates:
                    if debug:
                        print(
                            f"  row={row_index} 命中单位={matched_company or '-'} "
                            f"但{column_name}未提取到英文根域名/IP: {target_value!r}"
                        )
                    continue

                sheet_hits += 1
                if debug:
                    print(
                        f"  row={row_index} 命中单位={matched_company or '-'} "
                        f"{column_name}={target_value!r} -> {candidates}"
                    )

                for target in candidates:
                    results.append(target)

            if debug:
                print(f"sheet {sheet.title} 扫描行数: {scanned_rows} 命中行数: {sheet_hits}")

    finally:
        workbook.close()

    return results


def write_targets(output_file: Path, targets: list[str]) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    output_file.write_text("\n".join(sort_targets(targets)) + "\n", encoding="utf-8")


def choose_file_dialog(title: str, initialdir: Path, filetypes):
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        path = filedialog.askopenfilename(
            title=title,
            initialdir=str(initialdir),
            filetypes=filetypes,
        )
    finally:
        root.destroy()
    return Path(path) if path else None


def choose_save_dialog(title: str, initialdir: Path, initialfile: str):
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        path = filedialog.asksaveasfilename(
            title=title,
            initialdir=str(initialdir),
            initialfile=initialfile,
            defaultextension=".txt",
            filetypes=[("Text", "*.txt"), ("All files", "*.*")],
        )
    finally:
        root.destroy()
    return Path(path) if path else None


def main():
    parser = argparse.ArgumentParser(description="Enscan XLSX 单独提取脚本")
    parser.add_argument("--xlsx", default=None, help="指定 XLSX 文件路径")
    parser.add_argument("--company", default=None, help="指定 gongsi.txt 路径")
    parser.add_argument("--output", default=None, help="输出 txt 路径")
    parser.add_argument("--debug", action="store_true", help="输出详细匹配日志")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx) if args.xlsx else find_latest_xlsx(ENSCAN_DIR)
    if not xlsx_path:
        xlsx_path = choose_file_dialog(
            "选择 Enscan XLSX",
            ENSCAN_DIR,
            [("Excel", "*.xlsx"), ("All files", "*.*")]
        )
    if not xlsx_path:
        print("未选择 XLSX")
        raise SystemExit(1)

    company_path = Path(args.company) if args.company else DEFAULT_COMPANY_FILE
    if not company_path.exists():
        company_path = choose_file_dialog(
            "选择 gongsi.txt",
            DEFAULT_COMPANY_FILE.parent,
            [("Text", "*.txt"), ("All files", "*.*")]
        )
    if not company_path:
        print("未选择 gongsi.txt")
        raise SystemExit(1)

    output_path = Path(args.output) if args.output else DEFAULT_OUTPUT_FILE
    if args.output is None and not output_path:
        output_path = choose_save_dialog(
            "选择输出文件",
            PROJECT_ROOT,
            DEFAULT_OUTPUT_FILE.name,
        )
    if not output_path:
        print("未选择输出文件")
        raise SystemExit(1)

    if not company_path.exists():
        print(f"公司名单不存在: {company_path}")
        raise SystemExit(1)

    targets = sort_targets(
        extract_targets_from_xlsx_file(xlsx_path, company_path, debug=args.debug)
    )
    if not targets:
        print("没有提取到任何根域名")
        print("建议检查：")
        print("1. xlsx 是否真有数据行")
        print("2. 第二个 sheet 或其他 sheet 的整行是否能命中 gongsi.txt 的单位名称")
        print("3. 第三列是否确实是根域名")
        raise SystemExit(2)

    write_targets(output_path, targets)
    print(f"来源: {xlsx_path}")
    print(f"输出: {output_path}")
    print(f"提取数量: {len(targets)}")
    print("提取结果:")
    for item in targets:
        print(item)


if __name__ == "__main__":
    main()
